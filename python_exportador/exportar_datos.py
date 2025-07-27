# exportar_datos.py
import pandas as pd
from conexion_sql import obtener_conexion

# Consulta JOIN para obtener ventas con cliente y producto
consulta = """
SELECT 
    v.id_venta,
    v.id_fecha,
    c.nombre AS nombre_cliente,
    p.nombre_producto,
    v.cantidad,
    v.total_venta
FROM ventas v
JOIN clientes c ON v.id_cliente = c.id_cliente
JOIN productos p ON v.id_producto = p.id_producto
"""

# Ejecutar la consulta y guardar en DataFrame
conexion = obtener_conexion()
df = pd.read_sql(consulta, conexion)
conexion.close()

# Ver los primeros registros
print("\nPrimeros registros:")
print(df.head())

# Ver info de columnas (tipos de datos)
print("\nResumen del DataFrame:")
print(df.info())
# Convertir columna fecha a datetime
df['id_fecha'] = pd.to_datetime(df['id_fecha'], errors='coerce')

# Normalizar texto
df['nombre_cliente'] = df['nombre_cliente'].str.strip().str.title()
df['nombre_producto'] = df['nombre_producto'].str.strip().str.title()

# Eliminar duplicados si existieran
df.drop_duplicates(inplace=True)

# Crear columnas auxiliares para análisis
df['anio'] = df['id_fecha'].dt.year
df['mes'] = df['id_fecha'].dt.month

# Verificar resultados
print("\nDataFrame limpio:")
print(df.head())
print("\nResumen final:")
print(df.info())

# Ver si hay nulos
print("\nValores nulos por columna:")
print(df.isnull().sum())

print("\nResumen estadístico de columnas numéricas:")
print(df.describe())

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Ruta de salida
nombre_archivo = "ventas_limpias.xlsx"
ruta_archivo = f"datos/{nombre_archivo}"  # asegúrate de que la carpeta 'datos' exista

# Exportar a Excel con hoja "ventas_limpias"
with pd.ExcelWriter(ruta_archivo, engine="openpyxl", mode="w", datetime_format='YYYY-MM-DD') as writer:
    df.to_excel(writer, sheet_name="ventas_limpias", index=False)

print(f"\n✅ Archivo Excel exportado a {ruta_archivo}")

def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    col_final = chr(64 + max_col)  # Solo funciona hasta columna Z

    rango = f"A1:{col_final}{max_row}"
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    wb.save(ruta_archivo)
    print(f"✅ Tabla estructurada '{nombre_tabla}' creada en {ruta_archivo}")

# Llamar a la función para convertir en tabla estructurada
convertir_a_tabla_excel(ruta_archivo, "TablaVentas")

# Agrupar ventas por año y mes
ventas_mes = df.groupby(['anio', 'mes'])['total_venta'].sum().reset_index()

# Crear columna "Periodo" (ej: 2024-01)
ventas_mes['periodo'] = ventas_mes['anio'].astype(str) + '-' + ventas_mes['mes'].astype(str).str.zfill(2)

import matplotlib.pyplot as plt
import seaborn as sns
import os

# Gráfico 1: Ventas Totales por Mes
# Graficar
plt.figure(figsize=(10, 6))
sns.barplot(data=ventas_mes, x='periodo', y='total_venta', palette='Blues_d')

plt.title("Ventas Totales por Mes")
plt.xlabel("Mes")
plt.ylabel("Total Vendido")
plt.xticks(rotation=45)
plt.tight_layout()

# Guardar imagen
ruta_grafico = "graficas/ventas_por_mes.png"
plt.savefig(ruta_grafico)
plt.close()

print(f"📊 Gráfico guardado en: {ruta_grafico}")

top_productos = df.groupby("nombre_producto")["cantidad"].sum().nlargest(5).reset_index()

plt.figure(figsize=(8, 5))
sns.barplot(data=top_productos, x="cantidad", y="nombre_producto", palette="viridis")

plt.title("Top 5 Productos Más Vendidos")
plt.xlabel("Cantidad")
plt.ylabel("Producto")
plt.tight_layout()

plt.savefig("graficas/top5_productos.png")
plt.close()

print("📊 Gráfico Top 5 guardado en: graficas/top5_productos.png")

# Gráfico 2: Top 5 productos más vendidos
top_productos = df.groupby("nombre_producto")["cantidad"].sum().nlargest(5).reset_index()

plt.figure(figsize=(8, 5))
sns.barplot(data=top_productos, x="cantidad", y="nombre_producto", palette="viridis")

plt.title("Top 5 Productos Más Vendidos")
plt.xlabel("Cantidad")
plt.ylabel("Producto")
plt.tight_layout()

plt.savefig("graficas/top5_productos.png")
plt.close()

print("📊 Gráfico Top 5 guardado en: graficas/top5_productos.png")
# GRÁFICO TIPO TORTA (Pie Chart) – Participación por Cliente
# Agrupar total de ventas por cliente
ventas_cliente = df.groupby("nombre_cliente")["total_venta"].sum().sort_values(ascending=False)

# Si hay muchos clientes, mostramos solo los 6 primeros y agrupamos el resto como "Otros"
top_clientes = ventas_cliente.head(6)
otros = ventas_cliente[6:].sum()
top_clientes["Otros"] = otros

# Crear gráfico tipo torta
plt.figure(figsize=(7, 7))
plt.pie(top_clientes, labels=top_clientes.index, autopct="%1.1f%%", startangle=140)
plt.title("Participación de Clientes en Ventas")
plt.axis("equal")  # Círculo perfecto

# Guardar imagen
plt.savefig("graficas/participacion_clientes.png")
plt.close()

print("🥧 Gráfico de torta guardado como 'participacion_clientes.png'")

# import os
# from dotenv import load_dotenv

# load_dotenv()

# print("Servidor:", os.getenv("SQL_SERVER"))
# print("Base de datos:", os.getenv("SQL_DATABASE"))
# print("Usuario:", os.getenv("SQL_USER"))
# print("Contraseña:", os.getenv("SQL_PASSWORD"))
